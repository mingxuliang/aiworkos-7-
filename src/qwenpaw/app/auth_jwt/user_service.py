# -*- coding: utf-8 -*-
"""User CRUD service layer for JWT authentication.

All database operations are async and use SQLAlchemy sessions.
Password hashing uses bcrypt directly.
"""
from __future__ import annotations

import logging
from typing import Optional

import bcrypt
from sqlalchemy import select, delete, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from .models import User, Role, Permission, UserRole, RolePermission

logger = logging.getLogger(__name__)

# bcrypt for password hashing
def hash_password(password: str) -> str:
    """Hash a plaintext password using bcrypt."""
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain_password: str, hashed_password: str) -> bool:
    """Verify a plaintext password against a bcrypt hash."""
    return bcrypt.checkpw(
        plain_password.encode("utf-8"),
        hashed_password.encode("utf-8"),
    )


# ---------------------------------------------------------------------------
# User CRUD
# ---------------------------------------------------------------------------


async def create_user(
    session: AsyncSession,
    username: str,
    password: str,
    role_names: list[str] | None = None,
) -> User:
    """Create a new user with the given roles.

    Args:
        session: Async DB session.
        username: Unique username.
        password: Plaintext password (will be hashed).
        role_names: List of role names to assign (default: ["user"]).

    Returns:
        The created User instance.

    Raises:
        ValueError: If username already exists or role not found.
    """
    if role_names is None:
        role_names = ["user"]

    # Check uniqueness
    existing = await session.execute(
        select(User).where(User.username == username),
    )
    if existing.scalar_one_or_none() is not None:
        raise ValueError(f"Username '{username}' already exists")

    pw_hash = hash_password(password)
    user = User(username=username, password_hash=pw_hash)
    session.add(user)
    await session.flush()

    # Assign roles
    roles_result = await session.execute(
        select(Role).where(Role.name.in_(role_names)),
    )
    roles = roles_result.scalars().all()
    found_names = {r.name for r in roles}
    missing = set(role_names) - found_names
    if missing:
        raise ValueError(f"Roles not found: {missing}")

    for role in roles:
        session.add(UserRole(user_id=user.id, role_id=role.id))

    await session.commit()

    # Re-query with eager loading so that user.user_roles (and nested
    # role) are available in async context without triggering lazy-load.
    result = await session.execute(
        select(User)
        .where(User.id == user.id)
        .options(selectinload(User.user_roles).selectinload(UserRole.role)),
    )
    user = result.scalar_one()
    logger.info("User created: %s (roles=%s)", username, role_names)
    return user


async def authenticate_user(
    session: AsyncSession,
    username: str,
    password: str,
) -> Optional[User]:
    """Authenticate a user by username and password.

    Returns:
        User instance if valid, None otherwise.
    """
    result = await session.execute(
        select(User)
        .where(User.username == username, User.is_active.is_(True))
        .options(selectinload(User.user_roles).selectinload(UserRole.role)),
    )
    user = result.scalar_one_or_none()
    if user is None:
        return None
    if not verify_password(password, user.password_hash):
        return None
    return user


async def get_user_by_id(
    session: AsyncSession,
    user_id: int,
) -> Optional[User]:
    """Get a user by their database ID."""
    result = await session.execute(
        select(User)
        .where(User.id == user_id)
        .options(selectinload(User.user_roles).selectinload(UserRole.role)),
    )
    return result.scalar_one_or_none()


async def get_user_by_username(
    session: AsyncSession,
    username: str,
) -> Optional[User]:
    """Get a user by their username."""
    result = await session.execute(
        select(User)
        .where(User.username == username)
        .options(selectinload(User.user_roles).selectinload(UserRole.role)),
    )
    return result.scalar_one_or_none()


async def list_users(session: AsyncSession) -> list[User]:
    """List all users with their roles eagerly loaded."""
    result = await session.execute(
        select(User)
        .options(selectinload(User.user_roles).selectinload(UserRole.role))
        .order_by(User.id),
    )
    return list(result.scalars().all())


async def delete_user(session: AsyncSession, user_id: int) -> bool:
    """Delete a user by ID.

    Returns:
        True if a user was deleted.
    """
    user = await get_user_by_id(session, user_id)
    if user is None:
        return False
    # Cascade will handle UserRole deletions
    await session.delete(user)
    await session.commit()
    logger.info("User deleted: id=%d", user_id)
    return True


async def update_user_password(
    session: AsyncSession,
    user_id: int,
    new_password: str,
    new_password_repeat: str,
) -> tuple[bool, str]:
    """Update a user's password after verifying new passwords match.

    Returns:
        True on success, False if new passwords do not match.
    """
    user = await get_user_by_id(session, user_id)
    if user is None:
        return False, "User not found"
    if new_password != new_password_repeat:
        return False, "New passwords do not match"
    if verify_password(new_password, user.password_hash):
        return False, "New password is the same as the old password"
    user.password_hash = hash_password(new_password)
    await session.commit()
    logger.info("Password updated for user id=%d", user_id)
    return True, "Password updated successfully"


# ---------------------------------------------------------------------------
# Role management
# ---------------------------------------------------------------------------


async def assign_roles(
    session: AsyncSession,
    user_id: int,
    role_ids: list[int],
) -> bool:
    """Assign roles to a user (adds new, keeps existing).

    Returns:
        True on success.
    """
    # Get existing role assignments
    result = await session.execute(
        select(UserRole).where(UserRole.user_id == user_id),
    )
    existing_role_ids = {ur.role_id for ur in result.scalars().all()}

    for rid in role_ids:
        if rid not in existing_role_ids:
            session.add(UserRole(user_id=user_id, role_id=rid))

    await session.commit()
    logger.info("Roles assigned: user_id=%d role_ids=%s", user_id, role_ids)
    return True


async def remove_roles(
    session: AsyncSession,
    user_id: int,
    role_ids: list[int],
) -> bool:
    """Remove specific roles from a user.

    Returns:
        True on success.
    """
    for rid in role_ids:
        await session.execute(
            delete(UserRole).where(
                UserRole.user_id == user_id,
                UserRole.role_id == rid,
            ),
        )
    await session.commit()
    logger.info("Roles removed: user_id=%d role_ids=%s", user_id, role_ids)
    return True


async def list_roles(session: AsyncSession) -> list[Role]:
    """List all roles."""
    result = await session.execute(
        select(Role).options(
            selectinload(Role.role_permissions).selectinload(
                RolePermission.permission,
            ),
        ),
    )
    return list(result.scalars().all())


async def get_role_by_id(
    session: AsyncSession,
    role_id: int,
) -> Optional[Role]:
    """Get a role by its database ID."""
    result = await session.execute(
        select(Role)
        .where(Role.id == role_id)
        .options(
            selectinload(Role.role_permissions).selectinload(
                RolePermission.permission,
            ),
        ),
    )
    return result.scalar_one_or_none()


async def count_users_with_role(
    session: AsyncSession,
    role_id: int,
) -> int:
    """Count how many users are assigned to a given role."""
    result = await session.execute(
        select(func.count(UserRole.user_id)).where(
            UserRole.role_id == role_id,
        ),
    )
    return result.scalar_one()


async def create_role(
    session: AsyncSession,
    name: str,
    description: str = "",
) -> Role:
    """Create a new role.

    Args:
        session: Async DB session.
        name: Unique role name.
        description: Optional role description.

    Returns:
        The created Role instance.

    Raises:
        ValueError: If a role with the same name already exists.
    """
    existing = await session.execute(
        select(Role).where(Role.name == name),
    )
    if existing.scalar_one_or_none() is not None:
        raise ValueError(f"Role '{name}' already exists")

    role = Role(name=name, description=description)
    session.add(role)
    await session.commit()

    # Re-query with eager loading
    result = await session.execute(
        select(Role)
        .where(Role.id == role.id)
        .options(
            selectinload(Role.role_permissions).selectinload(
                RolePermission.permission,
            ),
        ),
    )
    role = result.scalar_one()
    logger.info("Role created: %s", name)
    return role


async def update_role(
    session: AsyncSession,
    role_id: int,
    name: Optional[str] = None,
    description: Optional[str] = None,
) -> Optional[Role]:
    """Update a role's name and/or description.

    Args:
        session: Async DB session.
        role_id: Role ID to update.
        name: New name (optional).
        description: New description (optional).

    Returns:
        Updated Role instance, or None if not found.

    Raises:
        ValueError: If the new name conflicts with an existing role.
    """
    role = await get_role_by_id(session, role_id)
    if role is None:
        return None

    if name is not None and name != role.name:
        # Check uniqueness of new name
        existing = await session.execute(
            select(Role).where(Role.name == name),
        )
        if existing.scalar_one_or_none() is not None:
            raise ValueError(f"Role '{name}' already exists")
        role.name = name

    if description is not None:
        role.description = description

    await session.commit()

    # Re-query with eager loading
    result = await session.execute(
        select(Role)
        .where(Role.id == role.id)
        .options(
            selectinload(Role.role_permissions).selectinload(
                RolePermission.permission,
            ),
        ),
    )
    role = result.scalar_one()
    logger.info("Role updated: id=%d name=%s", role_id, role.name)
    return role


async def delete_role(
    session: AsyncSession,
    role_id: int,
) -> tuple[bool, str]:
    """Delete a role by ID.

    Refuses to delete if any users are still assigned to this role.

    Returns:
        Tuple of (success, message).
    """
    role = await get_role_by_id(session, role_id)
    if role is None:
        return False, "Role not found"

    user_count = await count_users_with_role(session, role_id)
    if user_count > 0:
        return False, f"Cannot delete role '{role.name}': {user_count} user(s) are assigned to it"

    await session.delete(role)
    await session.commit()
    logger.info("Role deleted: id=%d name=%s", role_id, role.name)
    return True, "Role deleted"


async def list_permissions(session: AsyncSession) -> list[Permission]:
    """List all permissions."""
    result = await session.execute(select(Permission))
    return list(result.scalars().all())


# ---------------------------------------------------------------------------
# Permission checking
# ---------------------------------------------------------------------------


async def check_permission(
    session: AsyncSession,
    user_id: int,
    permission_code: str,
) -> bool:
    """Check if a user has a specific permission through any of their roles.

    Args:
        session: Async DB session.
        user_id: The user's database ID.
        permission_code: Permission code string (e.g. "agent:chat").

    Returns:
        True if the user has the permission.
    """
    # Get the user's role IDs
    result = await session.execute(
        select(UserRole.role_id).where(UserRole.user_id == user_id),
    )
    role_ids = [row[0] for row in result.all()]
    if not role_ids:
        return False

    # Get permission ID
    perm_result = await session.execute(
        select(Permission.id).where(Permission.code == permission_code),
    )
    perm_id = perm_result.scalar_one_or_none()
    if perm_id is None:
        return False

    # Check if any of the user's roles have this permission
    rp_result = await session.execute(
        select(RolePermission).where(
            RolePermission.role_id.in_(role_ids),
            RolePermission.permission_id == perm_id,
        ),
    )
    return rp_result.scalar_one_or_none() is not None


def user_has_role(user: User, role_name: str) -> bool:
    """Check if a loaded User object has a specific role.

    Uses the in-memory user_roles relationship (no DB query).
    """
    return any(ur.role.name == role_name for ur in user.user_roles)


# ---------------------------------------------------------------------------
# Paginated user list with search & filter
# ---------------------------------------------------------------------------


async def list_users_paginated(
    session: AsyncSession,
    page: int = 1,
    page_size: int = 10,
    username: str | None = None,
    role_name: str | None = None,
) -> tuple[list[User], int]:
    """Return a paginated list of users with optional filtering.

    Args:
        session: Async DB session.
        page: 1-based page number.
        page_size: Items per page.
        username: Optional username substring filter (case-insensitive).
        role_name: Optional role name filter.

    Returns:
        Tuple of (list of User, total_count).
    """
    # Base query with eager loading
    query = select(User).options(
        selectinload(User.user_roles).selectinload(UserRole.role),
    )
    count_query = select(func.count(User.id))

    # Apply filters
    if username:
        query = query.where(User.username.ilike(f"%{username}%"))
        count_query = count_query.where(User.username.ilike(f"%{username}%"))

    if role_name:
        query = (
            query
            .join(User.user_roles)
            .join(UserRole.role)
            .where(Role.name == role_name)
        )
        count_query = (
            count_query
            .join(User.user_roles)
            .join(UserRole.role)
            .where(Role.name == role_name)
        )

    # Get total count
    total = (await session.execute(count_query)).scalar_one()

    # Apply pagination
    offset = (page - 1) * page_size
    query = query.order_by(User.id).offset(offset).limit(page_size)

    result = await session.execute(query)
    users = list(result.scalars().unique().all())
    return users, total


# ---------------------------------------------------------------------------
# Batch delete users
# ---------------------------------------------------------------------------


async def batch_delete_users(
    session: AsyncSession,
    user_ids: list[int],
) -> int:
    """Delete multiple users by their IDs.

    Returns:
        Number of users actually deleted.
    """
    deleted = 0
    for uid in user_ids:
        user = await get_user_by_id(session, uid)
        if user is not None:
            await session.delete(user)
            deleted += 1
    await session.commit()
    logger.info("Batch deleted %d users: ids=%s", deleted, user_ids)
    return deleted


# ---------------------------------------------------------------------------
# Admin reset user password
# ---------------------------------------------------------------------------


async def reset_user_password(
    session: AsyncSession,
    user_id: int,
    new_password: str,
) -> tuple[bool, str]:
    """Admin resets a user's password without verifying the old one.

    Returns:
        Tuple of (success, message).
    """
    user = await get_user_by_id(session, user_id)
    if user is None:
        return False, "User not found"
    user.password_hash = hash_password(new_password)
    await session.commit()
    logger.info("Password reset by admin for user id=%d", user_id)
    return True, "Password reset successfully"


# ---------------------------------------------------------------------------
# Import users from Excel
# ---------------------------------------------------------------------------


async def import_users_from_excel(
    session: AsyncSession,
    file_bytes: bytes,
) -> tuple[int, list[str]]:
    """Parse an Excel file and create users in bulk.

    Expected columns: username, password, role (comma-separated role names).

    Returns:
        Tuple of (created_count, error_messages).
    """
    import io
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active

    created = 0
    errors: list[str] = []

    # Skip header row
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for idx, row in enumerate(rows, start=2):
        if not row or len(row) < 2:
            errors.append(f"Row {idx}: insufficient columns")
            continue

        username_val = str(row[0]).strip() if row[0] else ""
        password_val = str(row[1]).strip() if row[1] else ""
        role_val = str(row[2]).strip() if len(row) > 2 and row[2] else "user"

        if not username_val or not password_val:
            errors.append(f"Row {idx}: username or password is empty")
            continue

        role_names = [r.strip() for r in role_val.split(",") if r.strip()]
        if not role_names:
            role_names = ["user"]

        try:
            await create_user(session, username_val, password_val, role_names)
            created += 1
        except ValueError as exc:
            errors.append(f"Row {idx} ({username_val}): {exc}")

    wb.close()
    logger.info(
        "Excel import: created=%d, errors=%d", created, len(errors),
    )
    return created, errors
